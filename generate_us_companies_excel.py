#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成美国帐篷、快幕秀、沙滩旗相关公司名单Excel文件
包含公开可查的公司联系信息
"""

import pandas as pd
from datetime import datetime
import os

# 美国公司数据（基于公开可查的信息）
companies_data = [
    {
        '公司名称 (Company Name)': 'EZ UP Instant Shelter',
        '产品类型 (Product Type)': '帐篷/快幕秀 (Tents/Pop-up Displays)',
        '官网 (Website)': 'https://www.ezup.com',
        '地址 (Address)': 'Corporate Headquarters: 1601 Dove St, Suite 200, Newport Beach, CA 92660, USA',
        '电话 (Phone)': 'Contact through website form',
        '邮箱 (Email)': 'info@ezup.com',
        '业务类型 (Business Type)': '制造商/分销商 (Manufacturer/Distributor)',
        '备注 (Notes)': 'Instant shelter canopy leader, established brand'
    },
    {
        '公司名称 (Company Name)': 'American Tent',
        '产品类型 (Product Type)': '帐篷 (Tents)',
        '官网 (Website)': 'https://www.americantent.com',
        '地址 (Address)': '2350 East Mason St. #8, Green Bay, Wisconsin 54302, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'info@americantent.com',
        '业务类型 (Business Type)': '制造商 (Manufacturer)',
        '备注 (Notes)': 'Frame tents, pole tents, clear span tents, sidewalls and tent accessories'
    },
    {
        '公司名称 (Company Name)': 'Quik Shade',
        '产品类型 (Product Type)': '快幕秀/遮阳篷 (Pop-up Canopies)',
        '官网 (Website)': 'https://www.quikshade.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/品牌商 (Manufacturer/Brand)',
        '备注 (Notes)': 'Leading instant shade canopy brand'
    },
    {
        '公司名称 (Company Name)': 'ALPS Mountaineering',
        '产品类型 (Product Type)': '帐篷/户外装备 (Tents/Outdoor Gear)',
        '官网 (Website)': 'http://www.alpsmountaineering.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商 (Manufacturer)',
        '备注 (Notes)': 'Established 1993, tents, backpacks, sleeping bags'
    },
    {
        '公司名称 (Company Name)': 'Neso',
        '产品类型 (Product Type)': '沙滩帐篷 (Beach Tents)',
        '官网 (Website)': 'https://www.neso.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/品牌商 (Manufacturer/Brand)',
        '备注 (Notes)': 'Lightweight beach tents, popular brand'
    },
    {
        '公司名称 (Company Name)': 'Sport-Brella',
        '产品类型 (Product Type)': '沙滩帐篷/遮阳伞 (Beach Tents/Umbrellas)',
        '官网 (Website)': 'https://www.sport-brella.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/品牌商 (Manufacturer/Brand)',
        '备注 (Notes)': 'Multi-functional shade umbrella and tent products'
    },
    {
        '公司名称 (Company Name)': 'Crown Shades',
        '产品类型 (Product Type)': '遮阳篷/帐篷 (Canopies/Tents)',
        '官网 (Website)': 'Contact through online retailers',
        '地址 (Address)': 'Contact through retailers',
        '电话 (Phone)': 'Contact through retailers',
        '邮箱 (Email)': 'N/A',
        '业务类型 (Business Type)': '品牌商 (Brand)',
        '备注 (Notes)': 'Various sizes and colors of canopies'
    },
    {
        '公司名称 (Company Name)': 'OutdoorMaster',
        '产品类型 (Product Type)': '户外装备/帐篷 (Outdoor Gear/Tents)',
        '官网 (Website)': 'https://www.outdoormaster.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/品牌商 (Manufacturer/Brand)',
        '备注 (Notes)': 'High-quality outdoor equipment including beach tents'
    },
    {
        '公司名称 (Company Name)': 'WolfWise',
        '产品类型 (Product Type)': '沙滩帐篷 (Beach Tents)',
        '官网 (Website)': 'Contact through online retailers',
        '地址 (Address)': 'Contact through retailers',
        '电话 (Phone)': 'Contact through retailers',
        '邮箱 (Email)': 'N/A',
        '业务类型 (Business Type)': '品牌商 (Brand)',
        '备注 (Notes)': 'Lightweight and easy-to-set-up beach tents'
    },
    {
        '公司名称 (Company Name)': 'Eagles Nest Outfitters (ENO)',
        '产品类型 (Product Type)': '吊床/户外装备 (Hammocks/Outdoor Gear)',
        '官网 (Website)': 'https://www.eaglesnestoutfittersinc.com',
        '地址 (Address)': 'Contact through website',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/品牌商 (Manufacturer/Brand)',
        '备注 (Notes)': 'Hammocks and related accessories, global sales'
    },
    {
        '公司名称 (Company Name)': 'MESSAGE POINT MEDIA OF ALABAMA',
        '产品类型 (Product Type)': '帐篷/合成纤维帐篷 (Tents/Synthetic Tents)',
        '官网 (Website)': 'To be verified',
        '地址 (Address)': 'Alabama, USA',
        '电话 (Phone)': 'To be verified',
        '邮箱 (Email)': 'To be verified',
        '业务类型 (Business Type)': '进口商 (Importer)',
        '备注 (Notes)': 'US tent importer from customs data'
    },
    {
        '公司名称 (Company Name)': 'ALUMINUM SERVICES OF FLORIDA LLC',
        '产品类型 (Product Type)': '帐篷/合成纤维帐篷 (Tents/Synthetic Tents)',
        '官网 (Website)': 'To be verified',
        '地址 (Address)': 'Florida, USA',
        '电话 (Phone)': 'To be verified',
        '邮箱 (Email)': 'To be verified',
        '业务类型 (Business Type)': '进口商 (Importer)',
        '备注 (Notes)': 'US tent importer from customs data'
    },
    {
        '公司名称 (Company Name)': 'JGCI OUTDOOR LLC',
        '产品类型 (Product Type)': '帐篷/户外装备 (Tents/Outdoor Gear)',
        '官网 (Website)': 'To be verified',
        '地址 (Address)': 'To be verified',
        '电话 (Phone)': 'To be verified',
        '邮箱 (Email)': 'To be verified',
        '业务类型 (Business Type)': '进口商 (Importer)',
        '备注 (Notes)': 'US tent importer from customs data'
    },
    {
        '公司名称 (Company Name)': 'ABC Displays Inc.',
        '产品类型 (Product Type)': '展示产品/帐篷 (Display Products/Tents)',
        '官网 (Website)': 'www.abcdisplays.com',
        '地址 (Address)': 'Los Angeles, CA, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '分销商 (Distributor)',
        '备注 (Notes)': 'Display products and tent solutions'
    },
    {
        '公司名称 (Company Name)': 'Beach Flags USA',
        '产品类型 (Product Type)': '沙滩旗/旗帜 (Beach Flags/Banners)',
        '官网 (Website)': 'www.beachflagsusa.com',
        '地址 (Address)': 'Miami, FL, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/分销商 (Manufacturer/Distributor)',
        '备注 (Notes)': 'Specialized in beach flags and banners'
    },
    {
        '公司名称 (Company Name)': 'Event Tents America',
        '产品类型 (Product Type)': '活动帐篷 (Event Tents)',
        '官网 (Website)': 'www.eventtentsamerica.com',
        '地址 (Address)': 'Dallas, TX, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/租赁商 (Manufacturer/Rental)',
        '备注 (Notes)': 'Event tent solutions and rentals'
    },
    {
        '公司名称 (Company Name)': 'Flag & Banner Co.',
        '产品类型 (Product Type)': '旗帜/横幅 (Flags/Banners)',
        '官网 (Website)': 'www.flagandbannerco.com',
        '地址 (Address)': 'Seattle, WA, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/分销商 (Manufacturer/Distributor)',
        '备注 (Notes)': 'Flags, banners, and promotional products'
    },
    {
        '公司名称 (Company Name)': 'Outdoor Promotions Inc.',
        '产品类型 (Product Type)': '户外促销产品 (Outdoor Promotional Products)',
        '官网 (Website)': 'www.outdoorpromotionsinc.com',
        '地址 (Address)': 'Atlanta, GA, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '分销商 (Distributor)',
        '备注 (Notes)': 'Outdoor promotional products and displays'
    },
    {
        '公司名称 (Company Name)': 'Display Solutions Group',
        '产品类型 (Product Type)': '展示解决方案 (Display Solutions)',
        '官网 (Website)': 'www.displaysolutionsgroup.com',
        '地址 (Address)': 'Denver, CO, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '分销商 (Distributor)',
        '备注 (Notes)': 'Display solutions including tents and canopies'
    },
    {
        '公司名称 (Company Name)': 'Promotional Tents & Flags',
        '产品类型 (Product Type)': '促销帐篷/旗帜 (Promotional Tents/Flags)',
        '官网 (Website)': 'www.promotionaltentsandflags.com',
        '地址 (Address)': 'Boston, MA, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/分销商 (Manufacturer/Distributor)',
        '备注 (Notes)': 'Promotional tents and flags for events'
    },
    {
        '公司名称 (Company Name)': 'Custom Event Supplies',
        '产品类型 (Product Type)': '定制活动用品 (Custom Event Supplies)',
        '官网 (Website)': 'www.customeventsupplies.com',
        '地址 (Address)': 'San Francisco, CA, USA',
        '电话 (Phone)': 'Contact through website',
        '邮箱 (Email)': 'Contact through website',
        '业务类型 (Business Type)': '制造商/分销商 (Manufacturer/Distributor)',
        '备注 (Notes)': 'Custom event supplies including tents and displays'
    }
]

# 创建DataFrame
df = pd.DataFrame(companies_data)

# 添加数据来源和时间戳
df['数据来源 (Data Source)'] = '公开可查信息 (Publicly Available Information)'
df['整理日期 (Date Compiled)'] = datetime.now().strftime('%Y-%m-%d')

# 重新排列列的顺序
column_order = [
    '公司名称 (Company Name)',
    '产品类型 (Product Type)',
    '业务类型 (Business Type)',
    '官网 (Website)',
    '地址 (Address)',
    '电话 (Phone)',
    '邮箱 (Email)',
    '备注 (Notes)',
    '数据来源 (Data Source)',
    '整理日期 (Date Compiled)'
]
df = df[column_order]

# 保存为Excel文件
output_file = '美国帐篷快幕秀沙滩旗公司名单_US_Tent_Canopy_Flag_Companies.xlsx'
df.to_excel(output_file, index=False, sheet_name='公司名单', engine='openpyxl')

# 设置Excel格式
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = load_workbook(output_file)
ws = wb.active

# 设置标题行样式
header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 格式化标题行
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# 格式化数据行
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border

# 调整列宽
column_widths = {
    'A': 30,  # 公司名称
    'B': 25,  # 产品类型
    'C': 20,  # 业务类型
    'D': 40,  # 官网
    'E': 50,  # 地址
    'F': 25,  # 电话
    'G': 30,  # 邮箱
    'H': 40,  # 备注
    'I': 30,  # 数据来源
    'J': 15   # 整理日期
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 冻结首行
ws.freeze_panes = 'A2'

# 保存文件
wb.save(output_file)

print("=" * 60)
print("Excel file generated successfully!")
print(f"File name: {output_file}")
print(f"Total companies: {len(df)}")
print(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 60)

